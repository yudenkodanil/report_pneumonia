import pandas as pd
import config
import logging
from rich.console import Console
from rich.table import Table
from rich.logging import RichHandler
import copy
from openpyxl import load_workbook
from collections import defaultdict

# --- Настройка логирования ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[RichHandler(console=Console(), rich_tracebacks=True)]
)
logger = logging.getLogger("population_analysis")
console = Console()


# --- Утилиты ---
def increment_count(d: dict, key: str, val: int = 1):
    """Безопасно увеличивает счетчик в словаре"""
    d[key] = d.get(key, 0) + val


def get_age_group(age: float) -> str:
    for (low, high), label in zip(config.AGE_GROUP, config.AGE_GROUP_NAME):
        if label == config.AGE_GROUP_NAME[-1]:  # последняя группа "65 и старше"
            if age >= low:
                return label
        elif low <= age < high:
            return label
    return "Неизвестно"


def classify_status(status: str, age: float | None = None) -> str:
    """Классифицирует социальный статус через конфиг с проверкой возраста"""
    status = str(status).strip().lower()
    for group_name, keywords in config.SOCIAL_GROUPS.items():
        if status in [kw.lower() for kw in keywords]:
            if group_name in config.SOCIAL_GROUPS_ADULT_OVERRIDE and age is not None and age >= 18:
                return config.SOCIAL_GROUP_DEFAULT
            return group_name
    for keyword, group in config.KEYWORD_MAPPING.items():
        if keyword.lower() in status:
            if group in config.SOCIAL_GROUPS_ADULT_OVERRIDE and age is not None and age >= 18:
                return config.SOCIAL_GROUP_DEFAULT
            return group
    return config.SOCIAL_GROUP_DEFAULT

def classify_med_org(mo_name: str) -> str:
    """Классифицирует медорганизацию по словарю MED_ORGS"""
    if not isinstance(mo_name, str):
        return "Другие МО"

    name = mo_name.lower()
    for org, keywords in config.MED_ORGS.items():
        for kw in keywords:
            if kw in name:
                return org
    return "Другие МО"

def get_severity(place: str, hosp_date) -> dict:
    """Возвращает словарь с подсчетом тяжести заболевания с настройками из конфига"""
    severity = {key: 0 for key in config.SEVERITY_ORDER}
    place = str(place).strip()

    category = None
    for cat, places in config.SEVERITY_CATEGORIES.items():
        if place in places:
            category = cat
            break
    if category is None:
        category = config.SEVERITY_DEFAULT

    severity[f"{category}/всего"] = 1
    if pd.notna(hosp_date):
        severity[f"{category}/в т.ч. госпитализировано"] = 1

    return severity


# --- Предобработка файла ---
def preprocess_file(input_file: str) -> pd.DataFrame | None:
    try:
        df = pd.read_excel(input_file, header=None, skiprows=3, usecols=range(39))
        logger.info(f"Прочитано строк: {len(df)}, столбцов: {len(df.columns)}")

        if config.COLUMN_NAMES:
            df.columns = config.COLUMN_NAMES[:len(df.columns)]
        else:
            df.columns = [f"col_{i}" for i in range(1, len(df.columns) + 1)]

        df = df.loc[:, df.columns.notna()]

        # Заполняем пропуски в "Социальный статус"
        if config.COL_SOCIAL_STATUS in df.columns:
            idx_status = df.columns.get_loc(config.COL_SOCIAL_STATUS)
            df.iloc[:, idx_status] = df.iloc[:, idx_status].fillna(df.iloc[:, idx_status-1])

        # Преобразуем даты
        df[config.COL_BIRTH_DATE] = pd.to_datetime(df[config.COL_BIRTH_DATE], errors='coerce')
        df[config.COL_SUBMIT_DATE] = pd.to_datetime(df[config.COL_SUBMIT_DATE], errors='coerce')
        df['Возраст'] = (df[config.COL_SUBMIT_DATE] - df[config.COL_BIRTH_DATE]).dt.days / 365.25

        # --- Замена названий населённых пунктов на районы ---
        normalization = {
            "Сковородино": "Сковородинский район",
            # можно добавить другие соответствия при необходимости
        }
        if config.COL_DISTRICT in df.columns:
            df[config.COL_DISTRICT] = df[config.COL_DISTRICT].replace(normalization)

        # Экспортируем в новый Excel для проверки
        df.to_excel("df_filtred.xlsx", sheet_name="main", index=False)

        logger.info("Файл успешно предобработан")
        return df

    except FileNotFoundError:
        logger.error(f"Файл {input_file} не найден.")
        return None

# --- Создание новой структуры для региона ---
def new_region_structure():
    return {"age": {}, "social": {}, "severity": {}}

# --- Основной анализ ---
def analyze_population_full(df: pd.DataFrame) -> dict:
    """Считает возрастную структуру, соц. статус и степень тяжести заболевания"""
    result = {
        config.CITY_MAIN: new_region_structure(),
        "Районы": {}
    }

    mask_blg = (df[config.COL_DISTRICT] == config.CITY_MAIN) | \
               (df[config.COL_MED_ORG].isin(config.MED_ORG))

    for idx, row in df.iterrows():
        is_blg = mask_blg[idx]
        district = row[config.COL_DISTRICT] if not is_blg else config.CITY_MAIN

        age_group_label = get_age_group(row['Возраст'])
        social_group = classify_status(row.get(config.COL_SOCIAL_STATUS, ""), row['Возраст'])
        severity_counts = get_severity(row.get(config.COL_HOSP_PLACE, ""), row.get(config.COL_HOSP_DATE, pd.NaT))

        # --- создаем отдельную структуру для каждой территории ---
        target_dict = result[config.CITY_MAIN] if is_blg else result["Районы"].setdefault(
            district, new_region_structure()
        )

        increment_count(target_dict["age"], age_group_label)
        increment_count(target_dict["social"], social_group)
        for k, v in severity_counts.items():
            increment_count(target_dict["severity"], k, v)

    logger.info("Комплексный анализ завершен")
    return result

def analyze_by_med_org(df: pd.DataFrame):
    """Анализ по медицинским организациям города Благовещенска с учетом частичных совпадений названий МО"""
    
    results = defaultdict(lambda: {
        "age": defaultdict(int),
        "social": defaultdict(int),
        "severity": defaultdict(int),
    })

    col_age = "Возраст"  # уже подготовленная колонка
    col_social = config.COL_SOCIAL_STATUS
    col_hosp_place = config.COL_HOSP_PLACE
    col_hosp_date = config.COL_HOSP_DATE
    col_med_org = config.COL_MED_ORG
    city_main = config.CITY_MAIN

    # --- Маска для Благовещенска: либо город, либо МО из списка ---
    mask_blg = (
        (df[config.COL_DISTRICT] == city_main) |
        df[col_med_org].apply(lambda x: any(org.lower() in str(x).lower() for org in config.MED_ORGS.keys()))
    )

    df_blg = df[mask_blg]

    for _, row in df_blg.iterrows():
        # определяем МО с учетом частичного совпадения
        org = classify_med_org(row[col_med_org])

        # возраст
        age_group = get_age_group(row[col_age])
        results[org]["age"][age_group] += 1

        # социальный статус
        social_group = classify_status(row[col_social], row[col_age])
        results[org]["social"][social_group] += 1

        # степень тяжести
        sev = get_severity(row.get(col_hosp_place, ""), row.get(col_hosp_date))
        for key, val in sev.items():
            results[org]["severity"][key] += val

    return results

# --- Вывод таблиц ---
def _print_category_table(table: Table, title: str, data: dict, order: list[str], total_keys: list[str] | None = None):
    total = sum(data[k] for k in total_keys) if total_keys else sum(data.values())
    table.add_row(f"[bold blue]{title} ({total})[/bold blue]", "")
    for key in order:
        table.add_row(f"[green]{key}[/green]", f"[bright_yellow]{data.get(key, 0)}[/bright_yellow]")

def _print_table(title: str, data: dict):
    total = sum(data.get("age", {}).values())
    table = Table(title=f"{title} ({total} ЭИ)", title_style="bold magenta")
    table.add_column("Категория", style="cyan")
    table.add_column("Количество", style="yellow")

    _print_category_table(table, "Возрастная структура", data.get("age", {}), config.AGE_GROUP_NAME)
    _print_category_table(table, "Социальная структура", data.get("social", {}), config.SOCIAL_GROUP_ORDER)

    if "severity" in data:
        _print_category_table(
            table,
            "Степень тяжести",
            data.get("severity", {}),
            config.SEVERITY_ORDER,
            total_keys=config.SEVERITY_TOTAL_KEYS
        )

    console.print(table)

def print_structure(structure: dict):
    for region_name, data in structure.items():
        if region_name == "Районы":
            for district in config.ADM_TERR:
                if district in data:
                    _print_table(district, data[district])
        else:
            _print_table(region_name, data)

def fill_report(result: dict, template_file: str, output_file: str):
    wb = load_workbook(template_file)
    ws = wb.active

    # Cтрока для Благовещенска
    row = config.REPORT_LAYOUT["start_row"]

    # Cначала Благовещенск (как и было)
    _fill_block(ws, result[config.CITY_MAIN], config.REPORT_LAYOUT, row)
    row += 1

    # Затем все районы в порядке ADM_TERR
    for district in config.ADM_TERR:
        # !!! ВОТ ИСПРАВЛЕНИЕ !!!
        # Пропускаем Благовещенск, так как он уже обработан
        if district == config.CITY_MAIN:
            continue

        if district in result["Районы"]:
            _fill_block(ws, result["Районы"][district], config.REPORT_LAYOUT, row)
        else:
            # Eсли данных нет, вставляем нули
            empty_block = {
                "age": {k: 0 for k in config.AGE_GROUP_NAME},
                "social": {k: 0 for k in config.SOCIAL_GROUP_ORDER},
                "severity": {k: 0 for k in config.SEVERITY_ORDER}
            }
            _fill_block(ws, empty_block, config.REPORT_LAYOUT, row)
        
        row += 1 # Увеличиваем счетчик для каждого района (кроме пропущенного Благовещенска)
            
    wb.save(output_file)
    logger.info(f"Отчет по Амурской области сохранен в {output_file}")


def _fill_block(ws, data_block: dict, layout: dict, row: int):
    # Возрастные группы
    for age_group, col in layout["age"].items():
        ws.cell(row=row, column=col, value=data_block["age"].get(age_group, 0))

    # Социальные группы
    for social_group, col in layout["social"].items():
        ws.cell(row=row, column=col, value=data_block["social"].get(social_group, 0))

    # Степень тяжести
    for severity, col in layout["severity"].items():
        ws.cell(row=row, column=col, value=data_block["severity"].get(severity, 0))

def fill_report_by_med_org(result: dict, template_file: str, output_file: str):
    wb = load_workbook(template_file)
    ws = wb.active

    row = config.REPORT_LAYOUT["start_row"]  # начинаем с первой строки

    for org in config.MED_ORGS.keys():  # порядок фиксирован
        if org not in result:
            row += 1
            continue
        _fill_block(ws, result[org], config.REPORT_LAYOUT, row)
        row += 1

    wb.save(output_file)
    # print(f"Отчет по Благовещенску сохранен в {output_file}")
    logger.info(f"Отчет по Благовещенску сохранен в {output_file}")


# --- Главная функция ---
if __name__ == "__main__":
    logger.info("Программа запущена")
    df = preprocess_file(config.INPUT_FILE)
    if df is not None:
        # Анализ по территории (АО + районы)
        result_regions = analyze_population_full(df)
        print_structure(result_regions)
        fill_report(result_regions, config.TEMPLATE_FILE_AO, config.OUTPUT_FILE_AO)

        # Анализ по медицинским организациям
        result_med_orgs = analyze_by_med_org(df)
        fill_report_by_med_org(result_med_orgs, config.TEMPLATE_FILE_BLAG, config.OUTPUT_FILE_BLAG)
